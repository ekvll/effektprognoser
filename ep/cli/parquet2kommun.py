import numpy as np
import pandas as pd
import geopandas as gpd

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from tqdm import tqdm

from ep.cli.sql2parquet import (
    get_kommuner_in_region,
    parquet_filenames,
    load_parquet,
)
from ep.config import default_raps, default_years, EXCEL_DIR

"""
This script processes parquet files for a specific region. It generates Excel files with charts for effektbehov and elanvändning based on the data in the parquet files.

It performs the following steps:
1. Load parquet filenames for the specified region.
2. Get the list of kommuner in the region.
3. For each kommun, create DataFrames for effektbehov and elanvändning.
4. For each parquet file, load the data and filter it by the current kommun.
5. Aggregate the load profiles and calculate the maximum and sum for effektbehov and elanvändning.
6. Generate Excel filenames and file paths based on the kommun and year.
7. Save the DataFrames as Excel files with charts.
"""


def save_df_as_excel(df: pd.DataFrame | gpd.GeoDataFrame, filename: str) -> None:
    """
    Save a DataFrame or GeoDataFrame to an Excel file.

    Args:
        df (pd.DataFrame | gpd.GeoDataFrame): The DataFrame or GeoDataFrame to save.
        filename (str): The path where the Excel file will be saved.

    Returns:
        None

    Example:
        >>> df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})
        >>> save_df_as_excel(df, 'output.xlsx')
    """
    df.to_excel(filename, sheet_name="Data")


def gen_excel_filename(s1: str, s2: str, effektbehov: bool) -> str:
    """
    Generate an Excel filename based on the provided strings and effektbehov flag.

    Args:
        s1 (str): The first string to include in the filename.
        s2 (str): The second string to include in the filename.
        effektbehov (bool): Whether to include 'effektbehov' in the filename.

    Returns:
        str: The generated Excel filename.

    Example:
        >>> gen_excel_filename("1012", "Karlshamn", True)
        '1012_Karlshamn_effektbehov.xlsx'
    """
    filename = f"{s1}_{s2}"
    filename += "_effektbehov" if effektbehov else "_elanvandning"
    filename += ".xlsx"
    return filename


def gen_excel_filepath(region: str, filename: str) -> str:
    """
    Generate the full file path for an Excel file based on the region and filename.

    Args:
        region (str): The region for which the Excel file is created.
        filename (str): The name of the Excel file.

    Returns:
        str: The full file path for the Excel file.

    Example:
        >>> gen_excel_filepath("10", "kommun_2023_effektbehov.xlsx")
        'path/to/excel/10/kommun_2023_effektbehov.xlsx'
    """
    return EXCEL_DIR / region / filename


def make_excel_with_chart(
    df: pd.DataFrame | gpd.GeoDataFrame,
    filepath: str,
    effektbehov: bool,
) -> None:
    """
    Create an Excel file with a chart based on the provided DataFrame.

    Args:
        df (pd.DataFrame | gpd.GeoDataFrame): The DataFrame containing the data to be plotted.
        filepath (str): The path where the Excel file will be saved.
        effektbehov (bool): Whether to create a chart for effektbehov or elanvändning.

    Returns:
        None
    """
    df_transposed = df.transpose()

    save_df_as_excel(df_transposed, filepath)

    wb = load_workbook(filepath)
    ws = wb["Data"]

    chart = LineChart()
    chart.title = "Effektbehov per år" if effektbehov else "Elanvändning per år"
    chart.style = 1
    chart.y_axis.title = "Effektbehov (MW)" if effektbehov else "Elanvändning (MWh)"
    chart.x_axis.title = "År"

    # Reference data (each column = one line)
    # A1: header row; A2:A5 = years; B1: series name
    min_row = 1  # start of years
    max_row = ws.max_row
    min_col = 1  # first data column (original row names)
    max_col = ws.max_column

    data = Reference(
        ws, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row
    )
    categories = Reference(ws, min_col=1, max_col=1, min_row=min_row, max_row=max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    colors = [
        "FF0000",
        "00B050",
        "0070C0",
        "7030A0",
        "FFC000",
        "00FFFF",
        "FF00FF",
        "000000",
        "808080",
        "800000",
        "008080",
        "800080",
        "4682B4",
        "DAA520",
        "DC143C",
        "2E8B57",
        "A0522D",
        "5F9EA0",
        "D2691E",
        "9ACD32",
        "00008B",
        "B22222",
        "FF1493",
    ]

    for i, ser in enumerate(chart.series):
        ser.graphicalProperties.line.solidFill = colors[i % len(colors)]
    ws.add_chart(chart, "B10")
    wb.save(filepath)


def gen_array_of_zeros(length: int) -> np.ndarray:
    """
    Generate an array of zeros with the correct length based on the year.

    Args:
        length (int): The length of the array to generate.

    Returns:
        np.ndarray: An array of zeros with the specified length.

    Example:
        >>> gen_array_of_zeros(8760)
        array([0., 0., 0., ..., 0., 0., 0.])
    """
    return np.zeros(length)


def get_expected_length(year: int | str) -> int:
    """
    Get the expected length of the load profile based on the year.

    Args:
        year (int | str): The year for which to get the expected length.

    Returns:
        int: The expected length of the load profile array.

    Example:
        >>> get_expected_length(2023)
        8760
        >>> get_expected_length("2040")
        8784
    """
    if isinstance(year, int):
        year = str(year)

    return 8784 if year == "2040" else 8760


def verify_array_length(array: np.ndarray, expected_length: int) -> None:
    """
    Verify that the array has the expected length.

    Args:
        array (np.ndarray): The array to verify.
        expected_length (int): The expected length of the array.

    Returns:
        None

    Raises:
        ValueError: If the array does not have the expected length.

    Example:
        >>> verify_array_length(np.zeros(8760), 8760)
    """
    if array.shape[0] != expected_length:
        raise ValueError(
            f"Load profile has incorrect length: {array.shape[0]}, expected: {expected_length}"
        )


def get_year_and_raps(filename: str) -> tuple[str, str]:
    """
    Extract year and RAPS from the filename.

    Args:
        filename (str): The filename to extract from.

    Returns:
        tuple[str, str]: A tuple containing the year and RAPS.

    Example:
        >>> get_year_and_raps("kommun_2023_raps1_raps2.parquet")
        ('2023', 'raps1 raps2')
    """
    parts = filename.split("_")
    year = parts[1]
    raps = " ".join(parts[2:-1])
    return year, raps


def filter_df_by_kommun(
    gdf: pd.DataFrame | gpd.GeoDataFrame, kommun: str
) -> pd.DataFrame | gpd.GeoDataFrame:
    """
    Filter the GeoDataFrame by the specified kommun.

    Args:
        gdf (pd.DataFrame | gpd.GeoDataFrame): The GeoDataFrame to filter.
        kommun (str): The name of the kommun to filter by.

    Returns:
        pd.DataFrame | gpd.GeoDataFrame: A filtered GeoDataFrame containing only the specified kommun.

    Raises:
        ValueError: If the 'kn' column is not present in the GeoDataFrame.

    Example:
        >>> gdf = gpd.read_file("path/to/geojson")
        >>> filtered_gdf = filter_df_by_kommun(gdf, "Uppsala")
        >>> print(filtered_gdf)
    """
    if "kn" not in gdf.columns:
        raise ValueError("The GeoDataFrame does not contain a 'kn' column.")

    gdf_kommun = gdf[gdf["kn"] == kommun]

    if gdf_kommun.empty:
        # tqdm.write(f"No data found for kommun: {kommun}")
        return None

    return gdf_kommun


def gen_df_from_defaults(
    default_years: list[str], default_raps: list[str]
) -> pd.DataFrame:
    """
    Generate a DataFrame with default years and RAPS categories.

    Args:
        default_years (list[str]): List of years to use as columns.
        default_raps (list[str]): List of RAPS categories to use as index.

    Returns:
        pd.DataFrame: A DataFrame with the specified years as columns and RAPS categories as index.

    Example:
        >>> gen_df_from_defaults(['2023', '2040'], ['raps1', 'raps2'])
        Empty DataFrame
        Columns: [2023, 2040]
        Index: [raps1, raps2]
    """
    return pd.DataFrame(columns=default_years, index=default_raps)


def main(region: str) -> None:
    """Main function to process the region and generate Excel files with charts."""
    tqdm.write(f"Processing region: {region}")

    # Get the parquet filenames for the specified region
    filenames = parquet_filenames(region)

    # Get a list of kommuner found within the filenames
    kommuner = get_kommuner_in_region(filenames, region)

    tqdm.write(f"Found {len(kommuner['kommunnamn'])} kommuner in region {region}:")

    for kommun, kommunkod in zip(kommuner["kommunnamn"], kommuner["kommunkod"]):
        tqdm.write(f"Processing kommun: {kommun}")

        df_effektbehov = gen_df_from_defaults(default_years, default_raps)
        df_elanvandning = gen_df_from_defaults(default_years, default_raps)

        for filename in filenames:
            year, raps = get_year_and_raps(filename)

            gdf = load_parquet(filename, region)
            gdf_kommun = filter_df_by_kommun(gdf, kommun)

            if gdf_kommun is None:
                continue  # Skip if no data for the kommun

            expected_length = get_expected_length(year)
            aggregated_lp = gen_array_of_zeros(expected_length)

            for _, lp in enumerate(gdf_kommun["lp"]):
                lp_array = np.asarray(lp)
                verify_array_length(lp_array, expected_length)
                aggregated_lp += lp_array

            df_effektbehov.loc[raps, year] = max(aggregated_lp)
            df_elanvandning.loc[raps, year] = sum(aggregated_lp)

        # Generate the Excel filename
        filename_effektbehov = gen_excel_filename(
            str(kommunkod), kommun, effektbehov=True
        )
        filename_elanvandning = gen_excel_filename(
            str(kommunkod), kommun, effektbehov=False
        )

        # Generate the full file paths
        filepath_effektbehov = gen_excel_filepath(region, filename_effektbehov)
        filepath_elanvandning = gen_excel_filepath(region, filename_elanvandning)

        # Save the DataFrames as Excel files with charts
        make_excel_with_chart(df_effektbehov, filepath_effektbehov, effektbehov=True)
        make_excel_with_chart(df_elanvandning, filepath_elanvandning, effektbehov=False)


if __name__ == "__main__":
    tqdm.write("ep.cli.parquet2kommun")
    from ep.config import regions

    # regions = ["10"]
    for region in regions:
        main(region)
