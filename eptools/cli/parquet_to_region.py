import os
import sys
import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout
from eptools.utils.paths import EXCEL_DIR
from eptools.utils.groups import category_groups, custom_order
from eptools.processing.dataframe import (
    parquet_filenames,
    years_from_filenames,
    load_parquet,
)


def df_to_excel(df, region, calc):
    if calc == "max":
        excel_filename = f"{region} - Effektbehov (MW).xlsx"
    elif calc == "sum":
        excel_filename = f"{region} - Elanvandning (MWh).xlsx"
    # excel_filename = f"{region}.xlsx"
    path = os.path.join(EXCEL_DIR, region, excel_filename)

    df.to_excel(path, sheet_name="Data")
    # Load workbook
    wb = load_workbook(path)
    ws = wb["Data"]

    # Create Line Chart
    chart = LineChart()
    chart.title = "Effektbehov (MW)" if calc == "max" else "Elanvändning (MWh)"
    chart.y_axis.title = "Effektbehov (MW)" if calc == "max" else "Elanvändning (MWh)"
    chart.x_axis.title = "År"

    # Define X-axis categories (years)
    # Assuming the years start from column B (col=2) on row 1
    years = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=1)
    chart.set_categories(years)

    # Add one series per row (each Name)
    for i in range(2, 2 + len(df)):
        values = Reference(ws, min_col=2, max_col=5, min_row=i, max_row=i)
        name = ws.cell(row=i, column=1).value
        series = Series(values, title=name)
        chart.series.append(series)

    # Move legend to the right side, outside chart area
    # chart.legend.position = "r"  # Other options: 't' (top), 'b' (bottom), 'l' (left)
    # chart.legend.layout = Layout(manualLayout=ManualLayout(x=0.85, y=1, h=0.5, w=0.15))
    # Insert chart at a desired position
    ws.add_chart(chart, "G2")

    # Save workbook
    wb.save(path)
    # Load woorkbook and worksheet
    # wb = load_workbook(path)
    # ws = wb["Data"]

    # Create LineChart
    # chart = LineChart()
    # chart.title = "Effektbehov"
    # chart.style = 13
    # chart.y_axis.title = "Effektbehov"
    # chart.x_axis.title = "År"

    # Reference: vlaues (each row is a line)
    # We need to transpose the data for each row to become a line
    # for i, row in enumerate(df.itertuples(index=False), start=2):
    #    values = Reference(ws, min_col=2, max_col=ws.max_column, min_row=i, max_row=i)
    #    series = Series(values, title=f"{df.index[i - 2]}")
    #    chart.series.append(series)

    # Set X-axis labels
    # x_labels = Reference(ws, min_col=2, max_col=ws.max_column, min_row=1)
    # chart.set_categories(x_labels)

    # Add chart to worksheet
    # ws.add_chart(chart, "G2")

    # wb.save(path)


def filter_files(files, raps):
    raps = raps.replace(" ", "_")
    files_filtered = [f for f in files if raps in f]

    if "Flerbostadshus" in raps or "Smahus" in raps:
        if "LOM" not in raps:
            return sorted([f for f in files_filtered if "LOM" not in f])

    if "RAPS_8" in raps:
        if "8888" not in raps:
            return sorted([f for f in files_filtered if "8888" not in f])

    if "RAPS_7" in raps:
        if "7777" not in raps:
            return sorted([f for f in files_filtered if "7777" not in f])

    return sorted(files_filtered)


def run(region: str):
    print(f"Running pipeline for region {region}")

    files = parquet_filenames(region)
    years = years_from_filenames(files)

    for calc in ["max", "sum"]:
        df = pd.DataFrame(index=custom_order, columns=years)

        for raps in custom_order:
            files_filtered = filter_files(files, raps)

            if len(files_filtered) > len(years):
                raise ValueError(f"More files than years for {raps}: {files_filtered}")

            if len(files_filtered) < len(years) and len(files_filtered) != 0:
                print(f"Less files than years for {raps}: {files_filtered}")

            if len(files_filtered) == 0:
                print(f"No files for {raps}: {files_filtered}")
                continue

            years_filtered = years_from_filenames(files_filtered)
            for year in years_filtered:
                aggregated_lp = np.zeros(8784 if year == "2040" else 8760)
                files_year = [f for f in files_filtered if year in f]
                if len(files_year) != 1:
                    raise ValueError(f"More than one file for {year}: {files_year}")
                file = files_year[0]
                gdf = load_parquet(file, region)
                for lp in gdf["lp"]:
                    aggregated_lp += lp
                if calc == "max":
                    df.loc[raps, year] = float(np.max(aggregated_lp))
                elif calc == "sum":
                    df.loc[raps, year] = float(np.sum(aggregated_lp))
        df_to_excel(df, region, calc)


if __name__ == "__main__":
    regions = ["06", "07", "08", "10", "12", "13"]
    for region in regions:
        run(region)
        # postprocess(region)
